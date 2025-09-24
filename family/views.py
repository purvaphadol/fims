from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.http import JsonResponse
from .forms import FamilyHeadForm, HobbyFormSet, MemberFormset
from .models import FamilyHead, FamilyMember, Hobby, City, statusChoice
from django.http import FileResponse
import io, os
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from openpyxl import Workbook
from openpyxl.styles import *
import decimal

def home(request):
    return render(request, 'index.html')

def get_cities(request, state_id):
    # state_id = request.GET.get('state_id')
    cities = City.objects.filter(state_id=state_id).filter(status=statusChoice.ACTIVE).all()
    data = list(cities.values('id', 'city_name'))
    return JsonResponse(data, safe=False)

def family_form(request):
    head_form = FamilyHeadForm()
    hobby_formset = HobbyFormSet(prefix="hobbies")
    member_formset = MemberFormset(prefix="members")
    if request.method == 'POST':
        head_form = FamilyHeadForm(request.POST, request.FILES)
        hobby_formset = HobbyFormSet(request.POST, instance=head_form.instance, prefix="hobbies")
        member_formset = MemberFormset(request.POST, request.FILES, instance=head_form.instance, prefix="members")
        if head_form.is_valid() and hobby_formset.is_valid() and member_formset.is_valid():
            head = head_form.save()
            hobby_formset.instance = head
            hobby_formset.save()
            member_formset.instance = head
            member_formset.save()
            return JsonResponse({"success": True, "message": "Family Created Successfully."})
        else:
            return JsonResponse({
                "success": False,
                "head_errors": head_form.errors,
                "hobby_errors": hobby_formset.errors,
                "member_errors": member_formset.errors,
            }, status=400)
    context = {
        'head_form': head_form,
        'hobby_formset': hobby_formset,
        'member_formset': member_formset
    }
    return render(request, 'family_form.html', context)

def family_pdf(request, pk):
    head = FamilyHead.objects.get(pk=pk)
    members = FamilyMember.objects.filter(family_head=head).filter(status=statusChoice.ACTIVE)
    hobbies = Hobby.objects.filter(family_head=head).filter(status=statusChoice.ACTIVE)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{head.name}_family.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elements = []
    
    elements.append(Paragraph(f"Family Report: {head.surname} Family", styles['Heading1']))

    # Head details
    elements.append(Paragraph("Head Details", styles['Heading2']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Name: {head.name}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Surname: {head.surname}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Birth Date: {head.dob}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Mobile: {head.mobno}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Address: {head.address}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"State: {head.state.state_name}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"City: {head.city.city_name}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Pincode: {head.pincode}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Marital Status: {head.marital_status}", styles['Normal']))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Wedding Date: {head.wedding_date}", styles['Normal']))
    elements.append(Spacer(1, 4))

    # Head Photo
    if head.photo and hasattr(head.photo, 'path') and os.path.exists(head.photo.path):
        try:
            img = Image(head.photo.path, width=1.5*inch, height=2*inch)
            img.hAlign = 'CENTER'
            elements.append(Paragraph(f"Photo: ", styles['Normal']))
            elements.append(img)
            elements.append(Spacer(1, 12))
        except Exception as e:
            elements.append(Paragraph(f"Photo: {head.photo.path}", styles['Normal']))

    # Hobbies
    elements.append(Paragraph("Hobbies", styles['Heading3']))
    count=1
    for h in hobbies:
        elements.append(Paragraph(f"{count}. {h.hobby}", styles['Normal']))
        elements.append(Spacer(1, 4))

        count += 1
    elements.append(Spacer(1, 12))

    # Members
    elements.append(Paragraph("Members", styles['Heading2']))
    count = 1
    for m in members:
        elements.append(Paragraph(f"Member {count}", styles['Heading3']))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Name: {m.member_name}", styles['Normal']))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Birth Date: {m.member_dob}", styles['Normal']))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Marital Status: {m.member_marital}", styles['Normal']))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Wedding Date: {m.member_wedDate}", styles['Normal']))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Education: {m.education}", styles['Normal']))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Relation: {m.relation}", styles['Normal']))
        elements.append(Spacer(1, 4))

        # Member Photo
        if m.member_photo and hasattr(m.member_photo, 'path') and os.path.exists(m.member_photo.path):
            try:
                img = Image(m.member_photo.path, width=1.5*inch, height=2*inch)
                img.hAlign = 'CENTER'
                elements.append(Paragraph(f"Photo: ", styles['Normal']))
                elements.append(img)
                elements.append(Spacer(1, 12))
            except Exception as e:
                elements.append(Paragraph(f"Photo: {m.member_photo}", styles['Normal']))
  
        elements.append(Spacer(1, 12))
        count += 1

    doc.build(elements)
    return response

def family_excel(request, pk):
    head = FamilyHead.objects.get(id=pk)
    hobbies = Hobby.objects.filter(family_head=pk).filter(status=statusChoice.ACTIVE)
    members = FamilyMember.objects.filter(family_head=pk).filter(status=statusChoice.ACTIVE)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'family' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:L1')
    worksheet.merge_cells('A2:L2')
    first_cell = worksheet['A1']
    first_cell.value = "Family Report"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells('A3:L3')
    second_cell = worksheet['A3']
    second_cell.value = "Head Details"
    second_cell.font  = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Family Report'
    
    columns = ['Name', 'Surname', 'Birth Date', 'Mobile No', 'Address', 'State', 'City', 'Pincode', 'Marital Status', 'Wedding Date', 'Photo', 'Hobbies']
    worksheet.append(columns)

    hobby_list = []
    for hobby in hobbies:
        hobby_list.append(hobby.hobby)
    separator = ", "
    hobby_string = separator.join(hobby_list)

    worksheet.append([head.name, head.surname, str(head.dob), head.mobno, head.address, head.state.state_name, head.city.city_name, head.pincode, head.marital_status, str(head.wedding_date), str(head.photo), hobby_string])

    worksheet.merge_cells('A6:L6')
    worksheet.merge_cells('A7:G7')
    member_cell = worksheet['A7']
    member_cell.value = "Member Details"
    member_cell.font  = Font(bold=True, color="246ba1")
    member_cell.alignment = Alignment(horizontal="center", vertical="center")

    member_columns = ['Sr. No.', 'Name', 'Birth Date', 'Marital Status', 'Wedding Date', 'Education', 'Photo']
    worksheet.append(member_columns)
    count = 1
    for member in members:
        worksheet.append([count, member.member_name, str(member.member_dob), member.member_marital, str(member.member_wedDate), member.education, str(member.member_photo)])
        count += 1

    workbook.save(response)
    return response

def head_excel(request):
    heads = FamilyHead.objects.all().exclude(status=statusChoice.DELETE)

    if request.GET.get('search'):
        name = heads.filter(name__icontains=request.GET.get('search'))
        mobno = heads.filter(mobno__icontains=request.GET.get('search'))
        state = heads.filter(state__state_name__icontains=request.GET.get('search'))
        city = heads.filter(city__city_name__icontains=request.GET.get('search'))
        heads = name.union(mobno, state, city)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'all_family' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:P1')
    worksheet.merge_cells('A2:P2')
    first_cell = worksheet['A1']
    first_cell.value = "All Family Head Report"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'All Family Head Report'
    
    columns = ['Sr. No.', 'Member ID', 'Name', 'Surname', 'Birth Date', 'Mobile No', 'Address', 'State', 'City', 'Pincode', 'Marital Status', 'Wedding Date', 'Education', 'Relation', 'Photo', 'Hobbies', 'Head ID',]
    worksheet.append(columns)

    count = 1
    for head in heads:
        hobbies = Hobby.objects.filter(family_head=head.id).filter(status=statusChoice.ACTIVE)
        hobby_list = []
        for hobby in hobbies:
            hobby_list.append(hobby.hobby)
        separator = ", "
        hobby_string = separator.join(hobby_list)
        worksheet.append([count, "", head.name, head.surname, str(head.dob), head.mobno, head.address, head.state.state_name, head.city.city_name, head.pincode, head.marital_status, str(head.wedding_date), "", "Head", str(head.photo), hobby_string, head.id])
        members = FamilyMember.objects.filter(family_head=head.id).filter(status=statusChoice.ACTIVE)
        idx = 1
        for member in members:
            worksheet.append(["", idx, member.member_name, "", str(member.member_dob), "-", "", "", "", "", member.member_marital, str(member.member_wedDate), member.education, member.relation, str(member.member_photo), "", member.family_head.id])
            idx += 1
        count += 1

    workbook.save(response)
    return response

