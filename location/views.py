from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from family.models import State, City, statusChoice
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .forms import StateForm, CityForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import *
from django.http import JsonResponse
from django.template.loader import render_to_string

@login_required(login_url='login_page')
def state_list(request):
    states = State.objects.all().exclude(status=statusChoice.DELETE).order_by('-created_at')
    if request.GET.get('search'):
        states = states.filter(state_name__icontains=request.GET.get('search'))

    p = Paginator(states, 10)  
    page_number = request.GET.get('page')
    page_obj = p.get_page(page_number)
    totalPages = page_obj.paginator.num_pages
    context = {
        'page_obj': page_obj,
        'lastPage': totalPages,
        'totalPagelist': [n+1 for n in range(totalPages)],
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        state_html = render_to_string('state_list_template.html', context, request=request)
        return JsonResponse({'state_html': state_html})
    return render(request, 'state_list.html', context)

@login_required(login_url='login_page')
def create_state(request):
    state_form = StateForm()
    if request.method == 'POST':
        state_form = StateForm(request.POST)
        if state_form.is_valid():
            state_form.save()
            messages.success(request, 'State Created Successfully!')
            return redirect('state_list')
    context = {'state_form': state_form }
    return render(request, 'create_state.html', context)

@login_required(login_url='login_page')
def update_state(request, pk):
    state = State.objects.get(id=pk)
    state_form = StateForm(instance=state)
    if request.method == 'POST':
        state_form = StateForm(request.POST, instance=state)
        if state_form.is_valid():
            state_form.save()
            messages.success(request, 'State Updated Successfully!')
            return redirect('state_list')
    context = {'state_form': state_form }
    return render(request, 'update_state.html', context)

@login_required(login_url='login_page')
def delete_state(request, pk):
    state = State.objects.get(id=pk)
    state.status = statusChoice.DELETE
    state.save()
    City.objects.filter(state_id=state).update(status = statusChoice.DELETE)
    messages.success(request, 'State Deleted Successfully!')
    return redirect('state_list')

@login_required(login_url='login_page')
def city_list(request):
    cities = City.objects.all().exclude(status=statusChoice.DELETE).order_by('-created_at')
    if request.GET.get('search'):
        city = cities.filter(city_name__icontains=request.GET.get('search'))
        state = cities.filter(state__state_name__icontains=request.GET.get('search'))
        cities = city.union(state)

    p = Paginator(cities, 10)  
    page_number = request.GET.get('page')
    page_obj = p.get_page(page_number)
    totalPages = page_obj.paginator.num_pages
    context = {
        'page_obj': page_obj,
        'lastPage': totalPages,
        'totalPagelist': [n+1 for n in range(totalPages)],
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        city_html = render_to_string('city_list_template.html', context, request=request)
        return JsonResponse({'city_html': city_html})
    return render(request, 'city_list.html', context)

@login_required(login_url='login_page')
def create_city(request):
    city_form = CityForm()
    if request.method == 'POST':
        city_form = CityForm(request.POST)
        if city_form.is_valid():
            city_form.save()
            messages.success(request, 'City Created Successfully!')
            return redirect('city_list')
    context = {'city_form': city_form }
    return render(request, 'create_city.html', context)

@login_required(login_url='login_page')
def update_city(request, pk):
    city = City.objects.get(id=pk)
    city_form = CityForm(instance=city)
    if request.method == 'POST':
        city_form = CityForm(request.POST, instance=city)
        if city_form.is_valid():
            city_form.save()
            messages.success(request, 'City Updated Successfully!')
            return redirect('city_list')
    context = {'city_form': city_form }
    return render(request, 'update_city.html', context)

@login_required(login_url='login_page')
def delete_city(request, pk):
    city = City.objects.get(id=pk)
    city.status = statusChoice.DELETE
    city.save()
    messages.success(request, 'City Deleted Successfully!')
    return redirect('city_list')

def city_excel(request):
    cities = City.objects.all().exclude(status=statusChoice.DELETE)

    if request.GET.get('search'):
        city = cities.filter(city_name__icontains=request.GET.get('search'))
        state = cities.filter(state__state_name__icontains=request.GET.get('search'))
        cities = city.union(state)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'city' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')
    first_cell = worksheet['A1']
    first_cell.value = "City List"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'City'
    
    columns = ['ID', 'Name', 'State', 'Status']
    worksheet.append(columns)

    count = 1
    for city in cities:
        worksheet.append([count, city.city_name, city.state.state_name, city.status])
        count += 1

    workbook.save(response)
    return response

def state_excel(request):
    states = State.objects.all().exclude(status=statusChoice.DELETE)

    if request.GET.get('search'):
        states = states.filter(state_name__icontains=request.GET.get('search'))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'state' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:C1')
    worksheet.merge_cells('A2:C2')
    first_cell = worksheet['A1']
    first_cell.value = "State List"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'State'
    
    columns = ['ID', 'Name', 'Status']
    worksheet.append(columns)

    count = 1
    for state in states:
        worksheet.append([count, state.state_name, state.status])
        count += 1

    workbook.save(response)
    return response


# def delete_state(request, pk):
#     state = State.objects.get(id=pk)
#     if request.method == 'POST':
#         state.status = statusChoice.DELETE
#         state.save()
#         City.objects.filter(state_id=state).update(status = statusChoice.DELETE)
#         return redirect('state_list')
#     context = {'state':state}
#     return render(request, 'delete_state.html', context)